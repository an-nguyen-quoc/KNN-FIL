import openpyxl
from openpyxl.xml.constants import MIN_ROW
import pandas as pd 
import glob
from openpyxl import Workbook
import os

merge_row = 0
if not os.path.exists("merge_data_giatoc_z.xlsx"):
    merge_data = Workbook()
    merge_data.save('merge_data_giatoc_z.xlsx')
    
merge_data = openpyxl.load_workbook("merge_data_giatoc_z.xlsx")
merge_sheet = merge_data.active


for file in glob.glob(".\*\*.txt"):
    print(file)
    merge_row += 1
    with open(file, 'r') as infile: 
        df = pd.read_table(file,sep=',',header=None)
        df.to_excel(file.replace('.txt', '_all_feature.xlsx'), 'Sheet1', header=None, index=None, startrow=1, startcol=0)
        book = openpyxl.load_workbook(file.replace('.txt', '_all_feature.xlsx'))
        sheet = book.active
        row_index = 1
        # cell_index = 0
        previous_value = 0
        for row in sheet.iter_rows(min_row = 2):
            row_index += 1
            cell = sheet.cell(row=row_index, column = 8)
            previous_value = cell.value
            if row_index >= 5 and row_index <= 29:
                merge_sheet.cell(row = merge_row, column = row_index - 4, value = cell.value)
                     
        while row_index < 29:
            print("Here")
            row_index += 1
            merge_sheet.cell(row = merge_row, column = row_index - 4, value = previous_value)
        merge_sheet.cell(row = merge_row , column = 26, value = file[file.rindex("\\") - 1])       
        merge_data.save("merge_data_giatoc_z.xlsx")        
                
         
                
        
        
        
    