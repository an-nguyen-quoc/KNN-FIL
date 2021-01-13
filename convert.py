import openpyxl
from openpyxl.xml.constants import MIN_ROW
import pandas as pd 
import glob
from openpyxl import Workbook
import os

max_value = [500, 500, 500, 500, 500, 2.5, 12, 3]
merge_row = 0

for file in glob.glob(".\*\*.txt"):
    print(file)
    merge_row += 1
    with open(file, 'r') as infile: 
        df = pd.read_table(file,sep=',',header=None)
        df.to_excel(file.replace('.txt', '.xlsx'), 'Sheet1', header=None, index=None, startrow=1, startcol=0)
        book = openpyxl.load_workbook(file.replace('.txt', '.xlsx'))
        sheet = book.active
        if os.path.exists("merge_data.xlsx"):
            merge_data = openpyxl.load_workbook("merge_data.xlsx")
            merge_sheet = merge_data.active
            print("Code 1")
        else:
            print("Code 2")
            merge_data = Workbook()
            merge_sheet = merge_data.active
            for i in range (25):
                merge_sheet.cell(row = 1, column = i+1, value = i + 1)
            merge_sheet.cell(row = 1, column = 26, value = 'label')    
        #Them header
        sheet['A1'] = 'Ngon_1'
        sheet['B1'] = 'Ngon_2'
        sheet['C1'] = 'Ngon_3'
        sheet['D1'] = 'Ngon_4'
        sheet['E1'] = 'Ngon_5'
        sheet['F1'] = 'X'
        sheet['G1'] = 'Y'
        sheet['H1'] = 'Z'
        row_index = 1
        for row in sheet.iter_rows(min_row = 2):
            normal_value = 0
            row_index += 1
            for cell in row:
                #if float(cell.value) > max_value[cell.column - 1]:
                 #   cell.value = max_value[cell.column - 1]
                #print (row_index)
                # print (cell.value)
                normal_value = normal_value + float(cell.value) * float(cell.value) / (max_value[cell.column - 1] * max_value[cell.column - 1])
                if normal_value > 0:
                    previous_normal = normal_value
            sheet.cell(row = row_index, column = 9, value = normal_value)

            if row_index >= 5 and row_index <= 29:
                merge_sheet.cell(row = merge_row + 1, column = row_index - 4, value = normal_value)
        while row_index <= 29:
            row_index += 1
            print(previous_normal)
            merge_sheet.cell(row = merge_row + 1, column = row_index - 4, value = previous_normal)
        merge_sheet.cell(row = merge_row + 1, column = 26, value = file[file.rindex("\\") - 1])                              
        book.save(file.replace('.txt', '.xlsx'))
        merge_data.save('merge_data.xlsx')
                
        
        
        
    